# -*- coding: utf-8 -*-
"""生成《Avery 标准管理信息填写表单》的 xlsx 空白件 —— partner-docs-0728。

为什么要多一套 xlsx（合伙人原件是 docx）：
  docx 版每张表都写着「需要填写多条记录时，请复制本页」。一家 16 人的公司填表 01 就要
  复制 16 次页面，而 Avery 拿到的是 16 份互相重复的表格结构。xlsx 一张表一个 sheet、
  一行一条记录，对填表人和对 openpyxl 解析都是正确的形状。docx 原件照样发（好读、好打印、
  好在会上传阅），两者并列不是二选一。

设计约束（每一条都是拿真管线量出来的，不是审美）：
  * **单表头行，数据从第 2 行开始，表头里只有字段名。** 提示一律进批注：
    - 提示单独占一行会被 parse.py 的 xlsx 解析当成一条数据，凭空长出一个叫
      「工号，全表不可重复」的人；
    - 提示塞进表头单元格的中文括号里，会让 `_canon_header` 认不出这一列（它的兜底是
      「剥掉非汉字再查表」，中文括号跟着一起留下了）——实测抽出 0 人；
    - 表头单元格里的换行 `\n` 会把整条表头行切成好几段，碎片跟第一条真数据混行。
  * **列序 = 抽取器的读法。** 见 FORMS 里 01 表上方那段：姓名恒取第 0 列，
    role/team/tenure 有位置兜底，所以 姓名|岗位|部门|司龄 这个前四列是被约束死的。
  * **sheet 01 名字里带「名册」。** `parse.sniff_kind` 先看文件名、再看正文前 12 行，
    而 HeuristicExtractor 是**按 doc_kind 分支**的：判成 project 就永远不调
    `_people_from_roster`。把 roster 线索放进 sheet 名 = 放进正文，用户怎么重命名文件都还在。
    实测：判成 project → 0 人 + 1 个整文档的垃圾项目；判成 roster → 3 人全字段正确、零幽灵。
    （生产走 LLM 抽取，它只把 doc_kind 当提示词、不分支；这条只保 LLM 不可用时的兜底路径。）
  * **枚举列一律下拉。** 这是 xlsx 相对 docx 最大的一处真收益：它让用户填出来的词恰好
    落在 Avery 读得懂的词表里，而不是自由发挥。
  * **说明页带「Avery 现在吃到哪一层」一列。** 逐张说实话，见 INTAKE。03/06/07 目前只进
    材料库、不成独立卡片——不写清楚，用户填完一下午会以为传失败了。
  * **07 表的红线后果写死在说明页。** 实测：按「填写要点」填的 07 全部通过；一旦写进
    分数 / 百分比 / 等级 / 排名，`avery/ingest/pipeline.py:130` 是**整发上传硬拒**，
    不是丢掉那一格——同一发里的 01-06 会被一起废掉。.docx 原件只说了「本表不产生任何
    分数」，没说后果，补上。

重新生成：  python scripts/make-intake-xlsx.py
输出：      public/paperwork/forms/avery-intake-forms.xlsx
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUT = Path(__file__).resolve().parents[1] / "public" / "paperwork" / "forms" / "avery-intake-forms.xlsx"

INK = "2B2B28"
HEAD_FILL = PatternFill("solid", fgColor="EDE9E1")
NOTE_FILL = PatternFill("solid", fgColor="FBF8F2")
WARN_FILL = PatternFill("solid", fgColor="FBEFE8")
THIN = Side(style="thin", color="D8D2C6")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

H1 = Font(name="Microsoft YaHei", size=15, bold=True, color=INK)
H2 = Font(name="Microsoft YaHei", size=11, bold=True, color=INK)
BODY = Font(name="Microsoft YaHei", size=10, color=INK)
FAINT = Font(name="Microsoft YaHei", size=10, color="6B665C")

# 每张表：(sheet 名, 一句话用途, [(表头, 列宽, 批注 or None), ...], {列号: [下拉选项]})
#
# 🔴 表头单元格里**不许出现额外的汉字**（提示一律进批注）。
# avery/ingest/extract.py 的 `_canon_header` 把表头单元格映射到抽取器真正读的键
# （姓名→name / 岗位→role / 部门→team / 司龄→tenure / 主要负责→owns）。它的兜底是
# 「剥掉所有非汉字再查表」——所以 ASCII 的 ` *` 无害（`姓名 *` → `姓名` 命中），但
# 中文括号致命（`姓名（仅贵司内部留存）` → `姓名仅贵司内部留存` 查不到，整列失联）。
# 实测：带括号的版本从填好的样本里抽出 0 人。
FORMS: list[tuple[str, str, list[tuple[str, int, str | None]], dict[int, list[str]]]] = [
    (
        "01 组织与人员名册",
        "建立组织范围、汇报关系与职责边界。每位纳入范围的员工一行。",
        # 🔴 列序是被约束死的，不是排版偏好。`_people_from_roster` 有两条读法：
        #   ① 表头映射（col.get("role"/"team"/"tenure"/"owns")）
        #   ② 位置兜底（cells[1]=role / cells[2]=team / cells[3]=tenure），表头没说话时才用
        # 而**姓名恒取 cells[0]**，没有表头可言——所以姓名必须是第一列，否则
        # `_looks_like_name("SY-001")` 为假，整张表一个人都抽不出来（实测过）。
        # 后三列按 岗位/部门/司龄 排，是为了让①和②指向同一个格：任何一条路读出来的
        # 人卡都一样。把「入职日期」放到 cells[3] 会让位置兜底把一个日期塞进 tenure
        # （司龄是时长不是日期，extract.py:586 明写这两者不是一回事）。
        [
            ("姓名 *", 14,
             "必须是第一列，Avery 靠它认人。按数据处理协议约定，姓名与工号的对照关系由贵司保管；"
             "若贵司选择不提供姓名，本表就只会进材料库、不会长出人卡（见「00 读我」）。"),
            ("岗位 *", 18, "岗位名称，非职级。"),
            ("部门 *", 18, "完整部门名称。"),
            ("司龄", 12,
             "如「2 年」「8 个月」，选填。这一栏合伙人的 .docx 原件里没有，是表格版补的——"
             "有了它，人卡上才有「在这家公司多久了」这条上下文。"),
            ("主要负责 *", 46,
             "对应 .docx 原件表 01 的「核心职责」。写主要负责什么就行，不用罗列全部工作内容。"
             "例如「负责华东区渠道投放的方案与执行，对投放 ROI 负责」。多项之间用「、」或分号隔开。"),
            ("人员ID *", 16,
             "工号，全表不可重复。建议直接用贵司现有工号，例如 MKT-001。"
             "这个编号在其他所有表里都要用到，一旦定了就不要改。"),
            ("直属上级ID", 16, "须为本表已有的工号。"),
            ("任职状态 *", 14, None),
            ("入职日期", 16, "一律写成 2026-07-24 这样的格式，不要写「7月底」「下周」「近期」。"),
        ],
        {8: ["在职", "试用期", "待离职"]},
    ),
    (
        "02 项目台账",
        "统一项目名称、负责人、阶段与时间边界。每个在管项目一行。",
        [
            ("项目ID *", 20, "如 PRJ-2026-01，全表不可重复。"),
            ("项目名称 *", 26,
             "与内部通用叫法一致——请用团队日常沟通时的叫法，不要用立项文件里的全称，"
             "否则后续更新记录对不上。"),
            ("负责人ID *", 18, "须来自 01 表的工号。"),
            ("当前状态 *", 14,
             "Avery 读得懂「进行中」「已暂停」「已完成」。「未开始」是计划态、不是健康度，"
             "Avery 会如实留空并把卡片归到「状态未知」——这是有意的，不是没读到。"),
            ("开始日期 *", 16, "YYYY-MM-DD。"),
            ("计划完成日期 *", 18, "YYYY-MM-DD。"),
            ("完成进度 *", 16, "0-100 的整数，只填数字不带百分号。按贵司现有口径填即可，只要前后一致。"),
            ("项目目标 *", 46,
             "这个项目要达成什么，尽量写成可以判断达成与否的表述。「提升品牌影响力」无法判断，"
             "「秋季新品发布会覆盖 3 家行业媒体、留资 500 条」可以判断。"),
        ],
        {4: ["未开始", "进行中", "已暂停", "已完成"]},
    ),
    (
        "03 目标与指标",
        "明确被衡量对象、统计周期、目标与当前值。每个在跟踪的指标一行。",
        [
            ("指标ID *", 18, "如 KPI-001，全表不可重复。"),
            ("指标名称 *", 22, "如「渠道留资量」。"),
            ("关联对象 *", 14, None),
            ("关联对象ID *", 20, "对应的工号（来自 01 表）或项目ID（来自 02 表）。"),
            ("统计周期 *", 14, None),
            ("单位 *", 16,
             "如 万元、条、次、%。目标值和当前值只填数字，单位单独填在这一栏，否则系统无法比较。"),
            ("目标值 *", 14, "只填数字。"),
            ("当前值 *", 14, "只填数字。"),
            ("数据截止日 *", 18,
             "数据统计到哪一天。很重要——同一个指标不同人报的数可能截止到不同日期，"
             "不写清楚会得出错误结论。"),
            ("数据来源 *", 34,
             "哪个系统或哪份文件，必须具体到能查到的程度。「后台数据」不合格，"
             "「巨量引擎后台－秋季发布会计划－7月周报」合格。"),
        ],
        {3: ["人员", "项目", "部门"], 5: ["周", "月", "季度", "年"]},
    ),
    (
        "04 项目进度更新",
        "记录每次更新中已完成、下一步、阻塞事项与需要决策的事项。每次更新一行。",
        [
            ("更新ID *", 20, "如 UPD-20260724-01。"),
            ("项目ID *", 18, "须来自 02 表。"),
            ("更新日期 *", 16, "YYYY-MM-DD。"),
            ("责任人ID *", 18, "须来自 01 表。"),
            ("下次截止日", 16, "YYYY-MM-DD。"),
            ("本期完成 *", 40,
             "只写已经做完的事，正在做的写到「下期动作」里。这两栏混在一起，进度就失真了。"
             "避免「推进了」「跟进了」「基本完成」这类表述。"),
            ("下期动作 *", 40, "下一步具体做什么、谁做。"),
            ("当前阻塞", 40,
             "卡在哪里、卡了多久、在等谁。这是全表最有价值的一栏，也是唯一直接进项目卡的一栏"
             "——卡了多久、在等谁这两个信息一定要写，Avery 主要靠它判断风险。"),
            ("需管理者决策", 34, "需要经理拍板的事，没有就留空。"),
        ],
        {},
    ),
    (
        "05 风险与事项",
        "记录风险、阻塞、冲突、待决和客户反馈。每件事一行。",
        [
            ("事项ID *", 18, "如 ISS-001，全表不可重复。"),
            ("事项类型 *", 16,
             "只有「风险」「阻塞」两类会进项目卡（成为风险等级与阻塞项）；"
             "「冲突」「待决」「客户反馈」进材料库，Avery 回答时会引用。"),
            ("优先级 *", 12, None),
            ("关联项目ID", 18, "如与项目相关，须来自 02 表。"),
            ("发现日期 *", 16, "YYYY-MM-DD。"),
            ("处理截止日", 16, "YYYY-MM-DD。"),
            ("责任人ID", 18, "须来自 01 表。"),
            ("处理状态 *", 14, None),
            ("事实描述 *", 44,
             "只写已确认的事实，不写推测和评价，不要出现对人的评价。「小李配合度差」不合格，"
             "「主视觉终稿自 6 月 10 日起无更新，6 月 12 日群内说明在等设计定稿」合格。"
             "写冲突类事项时只记录发生了什么，不记录任何一方的动机判断。"),
            ("证据来源 *", 32,
             "哪份文件、哪次会议、哪条记录。不能空，也不能写「口头沟通」——必须能被第三方查到，"
             "例如「7 月 15 日项目周会纪要第 3 条」。"),
        ],
        {2: ["风险", "阻塞", "冲突", "待决", "客户反馈"], 3: ["高", "中", "低"],
         8: ["待处理", "处理中", "已关闭"]},
    ),
    (
        "06 周报与述职事实",
        "沉淀述职、自检和管理汇报中可核验的事实。每人每周期一行。",
        [
            ("记录ID *", 22, "如 RPT-20260724-001。"),
            ("人员ID *", 18, "须来自 01 表。"),
            ("述职周期 *", 20, "如 2026-W30 或 2026-07。"),
            ("提交日期 *", 16, "YYYY-MM-DD。"),
            ("已完成事实 *", 44,
             "具体做完的事，尽量带上数字和日期。「完成了 3 场直播，累计观看 1.2 万」"
             "比「直播工作顺利推进」有用得多。"),
            ("未达成及原因 *", 40,
             "哪些没做完、为什么。写客观情况，不用自我检讨——写清楚是资源不够、被别的事挤了，"
             "还是外部原因没到位。"),
            ("下一周期目标 *", 36, "下个周期要做完什么。"),
            ("需要支持", 34,
             "需要谁配合、需要什么资源。经常被留空，但这一栏最能帮到你——写下来，"
             "经理才知道该在哪儿使劲。"),
        ],
        {},
    ),
    (
        "07 评议与反馈",
        "记录主管或评议小组的事实反馈与改进约定，仅用于沟通、辅导与职责优化，不用于绩效评分。",
        [
            ("评议ID *", 20, "如 REV-2026Q3-001。"),
            ("被评议人员ID *", 20, "须来自 01 表。"),
            ("评议人ID *", 18, "须来自 01 表。"),
            ("评议周期 *", 16, "如 2026Q3。"),
            ("评议日期 *", 16, "YYYY-MM-DD。"),
            ("确认的优势 *", 40,
             "写具体做成的事，不写「能力强」「态度好」这类性格评价。"),
            ("需改进事项 *", 40,
             "写具体行为和场景，不写性格。「三次跨部门协调都等到周会才提出」是行为，"
             "「沟通主动性不足」是评价。⚠️ 本栏不要写分数、百分比、等级或排名——"
             "写了会导致整发上传被拒绝，详见「00 读我」。"),
            ("沟通后约定动作 *", 40,
             "双方谈完后定下来做什么，带具体时间点。必须是双方谈过之后共同确认的，"
             "这一栏是本表最重要的部分。"),
        ],
        {},
    ),
]

# 00 说明页的「Avery 现在吃到哪一层」——每一行都对着代码验过，不是估计。
INTAKE = {
    "01 组织与人员名册": "直接成人卡（姓名 / 部门 / 岗位 / 核心职责 / 汇报关系）",
    "02 项目台账": "直接成项目卡（状态 / 进度 / 目标 / 负责人）",
    "03 目标与指标": "进材料库，Avery 回答时会引用；暂不成独立指标卡",
    "04 项目进度更新": "「当前阻塞」进项目卡的阻塞项；其余进材料库",
    "05 风险与事项": "「风险」「阻塞」两类进项目卡；其余三类进材料库",
    "06 周报与述职事实": "进材料库；可能形成人身情境信号（只描述在扛什么，不评判）",
    "07 评议与反馈": "进材料库。见下方红线提醒",
}

WHEN = {
    "01 组织与人员名册": ("核心必填", "开始使用前，每位员工一行"),
    "02 项目台账": ("核心必填", "开始使用前，每个在管项目一行"),
    "03 目标与指标": ("核心必填", "开始使用前，每个在跟踪的指标一行"),
    "04 项目进度更新": ("建议补充", "每次项目有进展时填"),
    "05 风险与事项": ("建议补充", "发现风险、阻塞或待决事项时填"),
    "06 周报与述职事实": ("建议补充", "每个汇报周期填"),
    "07 评议与反馈": ("建议补充", "每次评议沟通后填"),
}


def _put(ws, row: int, col: int, value: str, *, font=BODY, wrap=True, fill=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = font
    c.alignment = Alignment(vertical="top", wrap_text=wrap)
    if fill is not None:
        c.fill = fill
    return c


def build_readme(wb: Workbook) -> None:
    # 🔴 说明页是**最后一个** sheet，不是第一个——而工作簿打开时仍然停在它上面（main() 里的
    # wb.active）。理由在解析侧：`_people_from_roster` 只看**整份文档里第一条含 `|` 的行**
    # 是不是花名册表头，说明页里的四列小表会抢占那个位置，于是 01 表的表头永远读不到，
    # 「主要负责」整列静默丢失（extract.py:1197 的 docstring 把这个失败态叫「一屏只有名字、
    # 底下什么都没有的团队页」）。实测：说明页在前 → 抽出 0 人；挪到最后 → 正常。
    ws = wb.create_sheet("00 读我")
    ws.sheet_view.showGridLines = False
    for col, width in zip("ABCDE", (22, 12, 30, 52, 2)):
        ws.column_dimensions[col].width = width

    r = 1
    _put(ws, r, 1, "Avery 标准管理信息 · 填写表单（表格版）", font=H1, wrap=False); r += 2

    _put(ws, r, 1,
         "这套表单是 Avery 的信息来源。请不要提交自由格式的文档、截图或聊天记录——"
         "表单之外的内容不会被采纳。\n"
         "本文件是《Avery 标准管理信息填写表单》(.docx) 的表格版：字段一致，"
         "改成一张表一个工作表、一行一条记录，不用再复制页面。两个版本填哪个都行。\n"
         "带 * 的是必填项。每一列的表头都挂了批注（单元格右上角的小红三角），"
         "鼠标悬停就能看到那一列该怎么填；更长的填写要点见 .docx 原件。")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    ws.row_dimensions[r].height = 62; r += 2

    _put(ws, r, 1, "三条基本要求", font=H2, wrap=False); r += 1
    for line in (
        "只写事实，不写判断 —— 写发生了什么、什么时候、有什么依据。不写对人的评价，不写推测原因。",
        "编号前后一致 —— 人员ID、项目ID、指标ID 一旦确定就不要改动。所有表靠这些编号相互关联。",
        "日期统一格式 —— 一律写成 2026-07-24 这样的格式，不要写「7月底」「下周」「近期」。",
        "不要动表头行和列的顺序 —— Avery 靠表头认列。尤其是 01 表：姓名必须留在第一列，"
        "挪走或删掉之后这张表只会进材料库，不会长出人卡。",
    ):
        _put(ws, r, 1, "· " + line)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        ws.row_dimensions[r].height = 18; r += 1
    r += 1

    _put(ws, r, 1, "七张表各管什么，以及 Avery 现在吃到哪一层", font=H2, wrap=False); r += 1
    _put(ws, r, 1, "填完这一整套之后 Avery 会长出什么，逐张说清楚。标「进材料库」的表不是白填——"
                   "Avery 回答问题时会检索并引用它们，只是暂时不会变成看板上的独立卡片。")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    ws.row_dimensions[r].height = 32; r += 1

    head = r
    for i, label in enumerate(("表", "是否必填", "什么时候填", "Avery 现在吃到哪一层"), start=1):
        c = _put(ws, head, i, label, font=H2, fill=HEAD_FILL)
        c.border = BORDER
    r += 1
    for name, _purpose, _cols, _dv in FORMS:
        must, when = WHEN[name]
        for i, val in enumerate((name, must, when, INTAKE[name]), start=1):
            c = _put(ws, r, i, val, fill=NOTE_FILL if i == 4 else None)
            c.border = BORDER
        ws.row_dimensions[r].height = 30
        r += 1
    r += 1

    _put(ws, r, 1, "表 07 请务必读这一段", font=H2, wrap=False); r += 1
    _put(ws, r, 1,
         "表 07 不产生任何分数或等级，Avery 也不提供绩效评分功能。这不只是一句承诺——它是"
         "系统层面的硬拦截：若「需改进事项」「确认的优势」等栏目里出现分数（如「绩效 2 分」）、"
         "百分比（如「完成度 82%」）、等级（如「绩效评级：不合格」）或排名（如「排名倒数第一」），"
         "整发上传会被拒绝，同一批里的其他表也会一起失败。\n"
         "按上面的「填写要点」写行为和场景就不会触发。已实测：「三次跨部门协调都等到周会才提出」"
         "「完成 3 场直播，累计观看 1.2 万」这类写法全部正常通过。",
         fill=WARN_FILL)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    ws.row_dimensions[r].height = 76; r += 2

    _put(ws, r, 1, "关于个人信息", font=H2, wrap=False); r += 1
    _put(ws, r, 1,
         "本套表单只采集与工作有关的信息。不采集身份证件号、生物识别信息、个人联系方式、"
         "家庭情况、简历、薪酬等信息，请不要在任何栏目中填写上述内容。\n"
         "表 01 的「姓名」列仅供贵司内部对照使用。若贵司选择不提供该列，删掉整列即可，"
         "其余各表靠工号照常关联。")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    ws.row_dimensions[r].height = 46; r += 2

    _put(ws, r, 1, "填写人（部门 / 工号）：", font=FAINT, wrap=False)
    _put(ws, r, 3, "填写日期：", font=FAINT, wrap=False)


def build_form(wb: Workbook, name, purpose, columns, dropdowns) -> None:
    ws = wb.create_sheet(name)
    ws.sheet_view.showGridLines = False

    for idx, (header, width, note) in enumerate(columns, start=1):
        letter = get_column_letter(idx)
        ws.column_dimensions[letter].width = width
        # 表头里的 \n 只是本文件的可读性排版，**绝不能写进单元格**：avery/ingest/parse.py 的
        # _parse_xlsx 把每个 sheet 按「一行一条、单元格用 | 分隔」压平，单元格内的换行会把
        # 表头行切成好几行（实测：「人员ID *」/「（工号…） | 姓名」…），于是表头碎片会和第
        # 一条真数据混在同一行喂给抽取器。视觉换行交给 wrap_text + 行高，那是纯显示层。
        c = ws.cell(row=1, column=idx, value=header.replace("\n", ""))
        c.font = H2
        c.fill = HEAD_FILL
        c.border = BORDER
        c.alignment = Alignment(vertical="center", wrap_text=True)
        if note:
            cm = Comment(note, "Avery")
            cm.width, cm.height = 320, 150
            c.comment = cm

    ws.row_dimensions[1].height = 44
    ws.freeze_panes = "A2"
    # 数据行留 200 行的边框与自动换行，够绝大多数团队一次填完；不够就照常往下加行。
    for row in range(2, 202):
        for idx in range(1, len(columns) + 1):
            c = ws.cell(row=row, column=idx)
            c.font = BODY
            c.alignment = Alignment(vertical="top", wrap_text=True)
            c.border = BORDER

    for col, options in dropdowns.items():
        dv = DataValidation(type="list", formula1='"' + ",".join(options) + '"', allow_blank=True)
        dv.error = "请从下拉列表中选择，Avery 按这套词表读取。"
        dv.errorTitle = "取值超出范围"
        ws.add_data_validation(dv)
        letter = get_column_letter(col)
        dv.add(f"{letter}2:{letter}201")

    # sheet 用途写进 tab 之外的地方会污染数据行，所以放批注挂在 A1 之外——用 sheet 属性存。
    ws.sheet_properties.tabColor = "D8D2C6"
    ws.title = name
    ws.sheet_state = "visible"
    ws.oddHeader.left.text = f"{name} —— {purpose}"


def main() -> None:
    wb = Workbook()
    wb.remove(wb.active)
    for name, purpose, columns, dropdowns in FORMS:
        build_form(wb, name, purpose, columns, dropdowns)
    build_readme(wb)                       # 见 build_readme 顶部：必须排在七张表之后
    wb.active = len(wb.sheetnames) - 1     # 但打开工作簿时停在说明页上
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"wrote {OUT}  ({OUT.stat().st_size} bytes, {len(wb.sheetnames)} sheets)")
    for s in wb.sheetnames:
        print(f"  · {s}")


if __name__ == "__main__":
    main()
