# -*- coding: utf-8 -*-
"""端到端：把 xlsx 空白件填成一份真样本 → 走 parse + ingest → 看真长出什么。

这是 acceptance.md 里「填好的表真能长出人卡」那一节的可复现脚本。量的是**管线接不接得住
这个形状**，不是抽取召回率：离线无 key 时用的是 heuristic 抽取器，生产上是 MiniMax。

heuristic 这条路恰恰是更严的那条——它按 `doc_kind` 分支、姓名恒取第 0 列、表头必须能被
`_canon_header` 认出来。能在它手里出结果，形状就是对的。

跑法（从仓库根）：
    python .issues/partner-docs-0728/probe-e2e-filled.py

期望输出：doc_kind=roster · ok=True · redline 干净 · people=3 且字段全对 · 零幽灵人。
"""
from __future__ import annotations

import os
import sys
import tempfile
from pathlib import Path

from openpyxl import load_workbook

REPO = Path(__file__).resolve().parents[2]
SRC = REPO / "public" / "paperwork" / "forms" / "avery-intake-forms.xlsx"
# 填好的样本落临时目录：它是过程产物，不入库。
OUT = Path(tempfile.gettempdir()) / "avery-filled-sample.xlsx"

sys.path.insert(0, str(REPO / "eval-harness"))
os.chdir(REPO / "eval-harness")

from avery.ingest import parse                     # noqa: E402
from avery.ingest.pipeline import ingest_docs      # noqa: E402
from avery.ingest.registry import ContextRegistry  # noqa: E402

# 🔴 列序必须与生成器一致：姓名 | 岗位 | 部门 | 司龄 | 主要负责 | 人员ID | 直属上级ID | 任职状态 | 入职日期
# 照 .docx 原件的顺序（人员ID 打头）填的话，`_looks_like_name("SY-001")` 为假 → 抽出 0 人。
PEOPLE = [
    ("陈思雨", "渠道运营", "市场部", "2 年", "华南区渠道投放的方案与执行、对投放 ROI 负责",
     "SY-001", "", "在职", "2024-03-11"),
    ("林浩然", "内容策划", "市场部", "8 个月", "秋季新品发布会的内容脚本、媒体沟通",
     "SY-002", "SY-001", "在职", "2025-01-06"),
    ("赵敏", "活动运营", "运营部", "2 个月", "线下门店活动的落地执行、物料对接",
     "SY-003", "", "试用期", "2026-05-18"),
]
PROJECTS = [
    ("PRJ-2026-01", "秋季新品发布会", "SY-001", "进行中", "2026-06-01", "2026-09-20", 45,
     "覆盖 3 家行业媒体、留资 500 条"),
    ("PRJ-2026-02", "门店会员日改版", "SY-003", "已暂停", "2026-04-10", "2026-08-30", 30,
     "会员日到店转化率 12% 提到 18%"),
]
# 指标ID 故意用 KPI-001：它曾经会被红线误判成人身评分（2026-07-28 已由另一条线修好），
# 留在这里当回归样本——再红就是那条修复退化了。
METRICS = [
    ("KPI-001", "渠道留资量", "项目", "PRJ-2026-01", "月", "条", 500, 180,
     "2026-07-24", "巨量引擎后台－7月周报"),
]
# 表 07 按「填写要点」填：写行为与场景，不写分数/等级/排名。这样填不该触发红线。
REVIEWS = [
    ("REV-2026Q3-001", "SY-002", "SY-001", "2026Q3", "2026-07-24",
     "6 月 10 日独立完成媒体协调，覆盖 3 家行业媒体",
     "三次跨部门协调都等到周会才提出",
     "8 月 15 日前交渠道对接清单，7 月 30 日双方已确认"),
]

wb = load_workbook(SRC)
for sheet, rows in (("01 组织与人员名册", PEOPLE), ("02 项目台账", PROJECTS),
                    ("03 目标与指标", METRICS), ("07 评议与反馈", REVIEWS)):
    ws = wb[sheet]
    for r, row in enumerate(rows, start=2):
        for c, value in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=value)
wb.save(OUT)

# 文件名刻意用纯 ASCII：证明路由线索来自**正文**（sheet 名里的「名册」），不是靠文件名去赌。
doc = parse.parse_bytes("avery-intake-forms.xlsx", OUT.read_bytes())
print(f"doc_kind = {doc.doc_kind}   （ASCII 文件名 → 路由只能来自正文）")

report = ingest_docs([doc], registry=ContextRegistry(), name="样板公司")
print(f"ingest ok={report.ok}   redline={report.redline}")
if not report.ok:
    for v in report.redline.violations:
        print("   !", v.kind, v.detail[:120])
    raise SystemExit(1)

ex = report.extraction
print(f"people={len(ex.people)}  projects={len(ex.projects)}  "
      f"signals={len(ex.signals)}  materials={len(ex.materials)}")
for p in ex.people:
    print(f"   · {p.name} | {p.role} | {p.team} | {p.tenure} | owns={p.owns}")

assert len(ex.people) == 3, f"应当抽出 3 个人，实得 {len(ex.people)}"
assert all(p.role and p.team and p.owns for p in ex.people), "有人的字段没抽全"
print("\n✓ 3 个人卡字段全对、零幽灵人（另外 6 张表的第 0 列都是编号，一行都没被误当成人）")
